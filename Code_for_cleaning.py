import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load the NASDAQ dataset
nasdaq_data = pd.read_csv(r'C:\Users\manik\Downloads\HistoricalData_1746126875835.csv')

# Load the NIFTY dataset
nifty_data = pd.read_csv(r'C:\Users\manik\Downloads\NIFTY 50_Historical_PR_01011990to01052025.csv')

# Ensure the Date columns are in datetime format
nasdaq_data['Date'] = pd.to_datetime(nasdaq_data['Date'])
nifty_data['Date'] = pd.to_datetime(nifty_data['Date'])

# Merge the datasets on the Date column
combined_data = pd.merge(nasdaq_data, nifty_data, on='Date', suffixes=('_NASDAQ', '_NIFTY'))

# Save the combined dataset to a new CSV file
combined_data.to_csv(r'C:\Users\manik\Downloads\combined_nasdaq_nifty.csv', index=False)

print("Datasets have been combined and saved to 'combined_nasdaq_nifty.csv'")

# Load the FTSE dataset from the Excel file, skipping the first 16 rows
ftse_data = pd.read_excel(r'C:\Users\manik\Downloads\FTSE Index values_1238.xlsx', skiprows=16)

# Remove the first column and keep the rest, excluding column 'C'
ftse_data_cleaned = ftse_data.drop(columns=['Unnamed: 0'])

# Define the path for the cleaned Excel file
cleaned_file_path = r'C:\Users\manik\Downloads\cleaned_ftse_data.xlsx'

# Save the cleaned FTSE dataset to the new Excel file
ftse_data_cleaned.to_excel(cleaned_file_path, index=False)

# Load the workbook and select the active worksheet
workbook = load_workbook(cleaned_file_path)
worksheet = workbook.active

# Set a fixed width for each column (e.g., width of 20)
for column_cells in worksheet.columns:
    column_letter = column_cells[0].column_letter  # Get the column name
    worksheet.column_dimensions[column_letter].width = 20  # Set the desired width

# Save the workbook with fixed column widths
workbook.save(cleaned_file_path)

print("Cleaned FTSE data has been saved to 'cleaned_ftse_data.xlsx' with fixed column widths.")

# File paths
dataset1_path = r'C:\Users\manik\Downloads\cleaned_ftse_data.xlsx'
dataset2_path = r'C:\Users\manik\Downloads\combined_nasdaq_nifty.csv'

# Load the datasets
nasdaq_data = pd.read_excel(dataset1_path)
nifty_data = pd.read_csv(dataset2_path)

# Convert the 'Date' column to datetime format
nasdaq_data['Date'] = pd.to_datetime(nasdaq_data['Date'])
nifty_data['Date'] = pd.to_datetime(nifty_data['Date'])

# Merge the datasets on the 'Date' column
combined_data = pd.merge(nasdaq_data, nifty_data, on='Date', suffixes=('_NASDAQ', '_NIFTY'))

# Save the combined dataset to a new CSV file
combined_data.to_csv(r'C:\Users\manik\Downloads\unified_dataset.csv', index=False)

print("Datasets combined successfully and saved as 'unified_dataset.csv'.")
